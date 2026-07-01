import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from .models import PayrollRun, PayrollEntry
from datetime import date
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from .services.dtr_client import DTRClient
from decimal import Decimal
from .services.payroll_calculator import _to_decimal
from audit.models import AuditLog
import io

from .models import Worker, TimesheetSync, PayrollRun, PayrollEntry, PayrollAdjustment
from .serializers import (
    WorkerSerializer,
    TimesheetSyncSerializer,
    PayrollRunSerializer,
    PayrollRunListSerializer,
    PayrollEntrySerializer,
    PayrollAdjustmentSerializer,
)
from .services.payroll_calculator import (
    sync_workers,
    sync_timesheets,
    build_payroll_run,
    recalculate_entry,
    approve_payroll_run,
)

from expenses.models import Transaction, Wallet

def get_client_ip(request):
    """Get client IP address from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sync_workers_view(request):
    """
    POST /api/payroll/workers/sync/
    Pulls all active workers from DTR and upserts into local Worker table.
    """
    try:
        workers = sync_workers()
        return Response({
            'status': 'synced',
            'count': len(workers)
        })
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_workers(request):
    """
    GET /api/payroll/workers/
    Returns all workers in the local Worker table.
    """
    workers = Worker.objects.filter(is_active=True).order_by('last_name')
    serializer = WorkerSerializer(workers, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sync_timesheets_view(request):
    """
    POST /api/payroll/timesheets/sync/
    Body: { period_start, period_end }
    Pulls raw timesheet data from DTR for the given period.
    This is Step 1 — no money calculated yet.
    """
    period_start = request.data.get('period_start')
    period_end = request.data.get('period_end')

    if not period_start or not period_end:
        return Response(
            {'error': 'period_start and period_end are required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        syncs = sync_timesheets(period_start, period_end)
        return Response({
            'status': 'synced',
            'period_start': period_start,
            'period_end': period_end,
            'records_synced': len(syncs),
        })
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_timesheets(request):
    """
    GET /api/payroll/timesheets/?period_start=&period_end=&worker_id=
    Returns synced timesheet records, optionally filtered.
    """
    period_start = request.query_params.get('period_start')
    period_end = request.query_params.get('period_end')
    worker_id = request.query_params.get('worker_id')

    syncs = TimesheetSync.objects.select_related('worker').all()

    if period_start:
        syncs = syncs.filter(period_start=period_start)
    if period_end:
        syncs = syncs.filter(period_end=period_end)
    if worker_id:
        syncs = syncs.filter(worker__id=worker_id)

    syncs = syncs.order_by('worker__last_name', 'date')
    serializer = TimesheetSyncSerializer(syncs, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_payroll_runs(request):
    """
    GET /api/payroll/runs/
    Returns all payroll runs — lightweight list view.
    Supports filter by status: ?status=draft or ?status=approved
    """
    runs = PayrollRun.objects.all().order_by('-generated_at')

    run_status = request.query_params.get('status')
    if run_status:
        runs = runs.filter(status=run_status)

    serializer = PayrollRunListSerializer(runs, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_payroll_run(request, pk):
    """
    GET /api/payroll/runs/<id>/
    Returns a single payroll run with all entries and adjustments.
    This is the Review Payroll screen data.
    """
    run = get_object_or_404(PayrollRun, pk=pk)
    serializer = PayrollRunSerializer(run)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def build_payroll_run_view(request):
    """
    POST /api/payroll/runs/build/
    Body: { period_start, period_end }
    Builds a draft payroll run from synced timesheets.
    This is Step 2 — calculates gross pay per worker.
    """
    period_start = request.data.get('period_start')
    period_end = request.data.get('period_end')

    if not period_start or not period_end:
        return Response(
            {'error': 'period_start and period_end are required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        payroll_run, entries = build_payroll_run(
            period_start,
            period_end,
            generated_by=request.user,
        )
        serializer = PayrollRunSerializer(payroll_run)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    except ValueError as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_payroll_run_view(request, pk):
    """
    POST /api/payroll/runs/<id>/approve/
    Body: { source_wallet }
    Finalizes the payroll run and logs to expenses.
    This is the 'Approve & Log to Expenses' button.
    """
    run = get_object_or_404(PayrollRun, pk=pk)

    if run.status == 'approved':
        return Response(
            {'error': 'This payroll run has already been approved'},
            status=status.HTTP_400_BAD_REQUEST
        )

    source_wallet = request.data.get('source_wallet')
    run = approve_payroll_run(run, source_wallet=source_wallet)
    serializer = PayrollRunSerializer(run)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_payroll_entry(request, pk):
    """
    GET /api/payroll/entries/<id>/
    Returns a single payroll entry with adjustments.
    This is the View Payroll Information modal data.
    """
    entry = get_object_or_404(
        PayrollEntry.objects.select_related('worker', 'payroll_run'),
        pk=pk
    )
    serializer = PayrollEntrySerializer(entry)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_adjustment(request, entry_id):
    """
    POST /api/payroll/entries/<entry_id>/adjustments/
    Body: { type, category, amount, reason }
    Adds a bonus or deduction to a payroll entry.
    This is the 'Add Adjustment' modal.
    """
    entry = get_object_or_404(PayrollEntry, pk=entry_id)

    if entry.payroll_run.status == 'approved':
        return Response(
            {'error': 'Cannot adjust an approved payroll run'},
            status=status.HTTP_400_BAD_REQUEST
        )

    serializer = PayrollAdjustmentSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(payroll_entry=entry)
        recalculate_entry(entry)
        updated_entry = PayrollEntrySerializer(entry)
        return Response(updated_entry.data, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_adjustment(request, entry_id, adjustment_id):
    """
    DELETE /api/payroll/entries/<entry_id>/adjustments/<adjustment_id>/
    Removes an adjustment and recalculates net pay.
    """
    entry = get_object_or_404(PayrollEntry, pk=entry_id)

    if entry.payroll_run.status == 'approved':
        return Response(
            {'error': 'Cannot adjust an approved payroll run'},
            status=status.HTTP_400_BAD_REQUEST
        )

    adjustment = get_object_or_404(PayrollAdjustment, pk=adjustment_id, payroll_entry=entry)
    adjustment.delete()
    recalculate_entry(entry)
    updated_entry = PayrollEntrySerializer(entry)
    return Response(updated_entry.data)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_worker_rates_view(request, pk):
    """
    PUT /api/payroll/workers/<id>/rates/
    Updates worker rates in DTR and syncs local Worker table.
    """
    worker = get_object_or_404(Worker, pk=pk)

    allowed_fields = [
        'hourly_rate',
        'daily_salary',
        'monthly_salary',
        'overtime_hourly_rate',
        'payment_type',
    ]

    data = {
        k: v for k, v in request.data.items()
        if k in allowed_fields
    }

    if not data:
        return Response(
            {'error': 'No valid fields provided'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        from .services.payroll_calculator import update_worker_rates
        worker = update_worker_rates(worker, data)
        serializer = WorkerSerializer(worker)
        return Response(serializer.data)
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def payroll_summary(request):
    period_start = request.query_params.get('period_start')
    period_end = request.query_params.get('period_end')

    if not period_start or not period_end:
        return Response(
            {'error': 'period_start and period_end are required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    workers = Worker.objects.filter(is_active=True)
    client = DTRClient()

    total_payroll = Decimal('0')
    total_employees = 0
    paid_employees = 0
    unpaid_employees = 0

    for worker in workers:
        all_shifts = TimesheetSync.objects.filter(
            worker=worker,
            date__gte=period_start,
            date__lte=period_end
        )

        if not all_shifts.exists():
            continue

        total_employees += 1

        # Calculate total payroll from ALL shifts in this period
        try:
            salary_data = client.get_salary(
                dtr_id=worker.dtr_id,
                start_date=period_start,
                end_date=period_end,
            )
            total_payroll += _to_decimal(salary_data.get('grossPay', 0))
        except Exception:
            pass

        # Check if all shifts are paid
        unpaid_shifts = all_shifts.filter(is_paid_in_dtr=False)
        if unpaid_shifts.exists():
            unpaid_employees += 1
        else:
            paid_employees += 1

    return Response({
        'period_start': period_start,
        'period_end': period_end,
        'total_payroll': total_payroll,
        'total_employees': total_employees,
        'paid_employees': paid_employees,
        'unpaid_employees': unpaid_employees,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_payroll_list(request):
    period_start = request.query_params.get('period_start')
    period_end = request.query_params.get('period_end')
    month = request.query_params.get('month')

    if not period_start or not period_end:
        if not month:
            return Response(
                {'error': 'period_start and period_end are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        year, month_num = map(int, month.split('-'))
        period_start = date(year, month_num, 1)
        if month_num == 12:
            period_end = date(year + 1, 1, 1)
        else:
            period_end = date(year, month_num + 1, 1)
    else:
        period_start = date.fromisoformat(period_start)
        period_end = date.fromisoformat(period_end)

    workers = Worker.objects.filter(is_active=True)
    client = DTRClient()
    result = []

    for worker in workers:
        try:
            # Get ALL shifts directly from DTR for this period
            timesheet = client.get_timesheet(
                start_date=str(period_start),
                end_date=str(period_end),
                employee_id=worker.dtr_id,
            )
            shifts = timesheet.get('data', [])
        except Exception:
            shifts = []

        if not shifts:
            continue

        # Read isPaid directly from DTR response
        total_hours = sum(s.get('totalHours', 0) for s in shifts)
        unpaid_shifts = [s for s in shifts if not s.get('isPaid', False)]
        all_paid = len(unpaid_shifts) == 0

        # Total gross for all shifts
        try:
            salary_data = client.get_salary(
                dtr_id=worker.dtr_id,
                start_date=str(period_start),
                end_date=str(period_end),
            )
            gross_pay = salary_data.get('grossPay', 0)
        except Exception:
            gross_pay = 0

        # Remaining gross for unpaid shifts only
        remaining_gross = 0
        if unpaid_shifts:
            unpaid_dates = sorted([s['date'] for s in unpaid_shifts])
            try:
                unpaid_salary = client.get_salary(
                    dtr_id=worker.dtr_id,
                    start_date=unpaid_dates[0],
                    end_date=unpaid_dates[-1],
                )
                remaining_gross = unpaid_salary.get('grossPay', 0)
            except Exception:
                pass

        payment_status = 'paid' if all_paid else 'unpaid'

        result.append({
            'id': worker.id,
            'employee_id': worker.employee_id,
            'full_name': worker.full_name,
            'position': worker.position or '—',
            'department': worker.department or '—',
            'payment_type': worker.payment_type,
            'hourly_rate': worker.hourly_rate,
            'monthly_salary': worker.monthly_salary,
            'total_hours': total_hours,
            'gross_pay': gross_pay,
            'remaining_gross': remaining_gross,
            'payment_status': payment_status,
        })

    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def workers_payment_status(request):
    """
    GET /api/payroll/workers/status/
    Returns worker payment status for a period.
    """
    period_start = request.query_params.get('period_start')
    period_end = request.query_params.get('period_end')

    if not period_start or not period_end:
        return Response(
            {'error': 'period_start and period_end are required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    workers = Worker.objects.filter(is_active=True)
    result = []

    for worker in workers:
        shifts = TimesheetSync.objects.filter(
            worker=worker,
            date__gte=period_start,
            date__lte=period_end
        )
        
        payroll_entries = PayrollEntry.objects.filter(
            worker=worker,
            payroll_run__period_start__gte=period_start,
            payroll_run__period_end__lte=period_end,
            payroll_run__status='approved'
        )
        
        total_hours = sum(s.total_hours for s in shifts)
        gross_pay = sum(e.gross_pay for e in payroll_entries)
        days_worked = shifts.count()

        result.append({
            'worker_id': worker.id,
            'dtr_id': worker.dtr_id,
            'employee_id': worker.employee_id,
            'full_name': worker.full_name,
            'position': worker.position,
            'department': worker.department,
            'payment_type': worker.payment_type,
            'hourly_rate': worker.hourly_rate,
            'monthly_salary': worker.monthly_salary,
            'gross_pay': gross_pay,
            'days_worked': days_worked,
            'total_hours': total_hours,
        })

    return Response(result)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def pay_employee(request, employee_id):
    period_start = request.data.get('period_start')
    period_end = request.data.get('period_end')
    wallet_id = request.data.get('wallet_id')
    bonus = Decimal(str(request.data.get('bonus', 0)))
    overtime_pay = Decimal(str(request.data.get('overtime', 0)))
    deduction = Decimal(str(request.data.get('deduction', 0)))
    note = request.data.get('note', '')

    if not period_start or not period_end:
        return Response(
            {'error': 'period_start and period_end are required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not wallet_id:
        return Response(
            {'error': 'wallet_id is required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        worker = Worker.objects.get(id=employee_id)
    except Worker.DoesNotExist:
        return Response(
            {'error': 'Worker not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    client = DTRClient()

    from datetime import date as date_type
    period_start_date = date_type.fromisoformat(period_start)
    period_end_date = date_type.fromisoformat(period_end)

    # Get ALL shifts in the period
    all_shifts = TimesheetSync.objects.filter(
        worker=worker,
        date__gte=period_start_date,
        date__lte=period_end_date,
    )

    # Filter only UNPAID shifts
    unpaid_shifts = all_shifts.filter(is_paid_in_dtr=False)

    if not unpaid_shifts.exists():
        return Response(
            {'error': 'This employee has already been paid for this period'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Use ONLY the date range of unpaid shifts
    unpaid_dates = unpaid_shifts.order_by('date')
    unpaid_start = str(unpaid_dates.first().date)
    unpaid_end = str(unpaid_dates.last().date)
    unpaid_shift_ids = list(
        unpaid_shifts.exclude(dtr_shift_id__isnull=True).values_list('dtr_shift_id', flat=True)
    )

    print(f" Unpaid shifts found: {unpaid_shifts.count()}")
    print(f" Unpaid date range: {unpaid_start} to {unpaid_end}")

    try:
        # Calculate salary ONLY for unpaid shift date range
        salary_data = client.get_salary(
            dtr_id=worker.dtr_id,
            start_date=unpaid_start,
            end_date=unpaid_end,
        )
        gross = Decimal(str(salary_data.get('grossPay', 0)))

        print(f" Gross pay for unpaid shifts: {gross}")

        # Find OR CREATE a single draft run for this period (NOT per employee)
        payroll_run, created = PayrollRun.objects.get_or_create(
            period_start=period_start_date,
            period_end=period_end_date,
            status='draft',
            defaults={
                'generated_by': request.user,
                'notes': note,
                'source_wallet': wallet_id,
            }
        )

        # Check if this employee already has an entry in the run
        existing_entry = PayrollEntry.objects.filter(
            payroll_run=payroll_run,
            worker=worker
        ).first()

        if existing_entry:
            # Update existing entry
            entry = existing_entry
            entry.total_working_hours = Decimal(str(salary_data.get('totalRegularHours', 0)))
            entry.total_overtime_hours = Decimal(str(salary_data.get('totalOvertimeHours', 0)))
            entry.total_days_worked = salary_data.get('daysWorked', 0)
            entry.payment_type = worker.payment_type
            entry.hourly_rate = worker.hourly_rate
            entry.overtime_hourly_rate = worker.overtime_hourly_rate
            entry.monthly_salary = worker.monthly_salary
            entry.gross_pay = gross
            entry.total_additions = Decimal('0')
            entry.total_deductions = Decimal('0')
            entry.net_pay = gross
            entry.save()
        else:
            # Create new entry for this employee
            entry = PayrollEntry.objects.create(
                payroll_run=payroll_run,
                worker=worker,
                total_working_hours=Decimal(str(salary_data.get('totalRegularHours', 0))),
                total_overtime_hours=Decimal(str(salary_data.get('totalOvertimeHours', 0))),
                total_days_worked=salary_data.get('daysWorked', 0),
                payment_type=worker.payment_type,
                hourly_rate=worker.hourly_rate,
                overtime_hourly_rate=worker.overtime_hourly_rate,
                monthly_salary=worker.monthly_salary,
                gross_pay=gross,
                total_additions=Decimal('0'),
                total_deductions=Decimal('0'),
                net_pay=gross,
            )

        # Add adjustments
        if bonus > 0:
            PayrollAdjustment.objects.create(
                payroll_entry=entry,
                type='addition',
                category='bonus',
                amount=bonus,
                reason='Bonus'
            )
        if overtime_pay > 0:
            PayrollAdjustment.objects.create(
                payroll_entry=entry,
                type='addition',
                category='overtime_premium',
                amount=overtime_pay,
                reason='Overtime'
            )
        if deduction > 0:
            PayrollAdjustment.objects.create(
                payroll_entry=entry,
                type='deduction',
                category='other',
                amount=deduction,
                reason='Deduction'
            )

        recalculate_entry(entry)

        # Update run totals
        all_entries = payroll_run.entries.all()
        payroll_run.total_gross_pay = sum(e.gross_pay for e in all_entries)
        payroll_run.total_deductions = sum(e.total_deductions for e in all_entries)
        payroll_run.total_net_pay = sum(e.net_pay for e in all_entries)
        payroll_run.save()

        approve_payroll_run(
            payroll_run,
            wallet_id=wallet_id,
            user=request.user,
            shift_ids=unpaid_shift_ids,
        )

        return Response({
            'status': 'paid',
            'worker': worker.full_name,
            'gross_pay': str(entry.gross_pay),
            'net_pay': str(entry.net_pay),
            'period_start': period_start,
            'period_end': period_end,
            'unpaid_shifts_paid': unpaid_shifts.count(),
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_payroll_xlsx(request):
    """
    GET /api/payroll/export/xlsx/?period_start=&period_end=
    Exports one row per employee — mirrors the payroll page UI.
    """
    period_start = request.query_params.get('period_start')
    period_end = request.query_params.get('period_end')

    if not period_start or not period_end:
        return Response({'error': 'period_start and period_end are required'}, status=400)

    period_start_date = date.fromisoformat(period_start)
    period_end_date = date.fromisoformat(period_end)

    workers = Worker.objects.filter(is_active=True)
    client = DTRClient()
    rows = []

    for worker in workers:
        try:
            timesheet = client.get_timesheet(
                start_date=str(period_start_date),
                end_date=str(period_end_date),
                employee_id=worker.dtr_id,
            )
            shifts = timesheet.get('data', [])
        except Exception:
            shifts = []

        if not shifts:
            continue

        unpaid_shifts = [s for s in shifts if not s.get('isPaid', False)]
        all_paid = len(unpaid_shifts) == 0
        total_hours = sum(s.get('totalHours', 0) for s in shifts)

        try:
            salary_data = client.get_salary(
                dtr_id=worker.dtr_id,
                start_date=str(period_start_date),
                end_date=str(period_end_date),
            )
            gross_pay = float(salary_data.get('grossPay', 0))
        except Exception:
            gross_pay = 0

        remaining_gross = 0
        if unpaid_shifts:
            unpaid_dates = sorted([s['date'] for s in unpaid_shifts])
            try:
                unpaid_salary = client.get_salary(
                    dtr_id=worker.dtr_id,
                    start_date=unpaid_dates[0],
                    end_date=unpaid_dates[-1],
                )
                remaining_gross = float(unpaid_salary.get('grossPay', 0))
            except Exception:
                pass

        # Sum all approved payroll entries for this worker in this period
        # to get actual total paid out (may differ from gross if adjustments)
        approved_entries = PayrollEntry.objects.filter(
            worker=worker,
            payroll_run__period_start__gte=period_start_date,
            payroll_run__period_end__lte=period_end_date,
            payroll_run__status='approved',
        )
        total_net_paid = sum(float(e.net_pay) for e in approved_entries)
        total_bonus = sum(
            float(a.amount)
            for e in approved_entries
            for a in e.adjustments.filter(type='addition', category='bonus')
        )
        total_overtime = sum(
            float(a.amount)
            for e in approved_entries
            for a in e.adjustments.filter(type='addition', category='overtime_premium')
        )
        total_deductions = sum(float(e.total_deductions) for e in approved_entries)

        rate = float(worker.hourly_rate or worker.monthly_salary or 0)
        rate_label = (
            f"₱{rate:,.2f}/hr" if worker.payment_type == 'hourly'
            else f"₱{rate:,.2f}/mo" if worker.payment_type == 'monthly'
            else '—'
        )

        rows.append({
            'employee_id': worker.employee_id,
            'full_name': worker.full_name,
            'position': worker.position or '—',
            'department': worker.department or '—',
            'payment_type': worker.payment_type,
            'rate_label': rate_label,
            'total_hours': total_hours,
            'gross_pay': gross_pay,
            'bonus': total_bonus,
            'overtime': total_overtime,
            'deductions': total_deductions,
            'amount_due': remaining_gross,
            'net_paid': total_net_paid,
            'status': 'Paid' if all_paid else 'Unpaid',
        })

    # --- Build XLSX ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"

    navy   = "1B2A4A"
    orange = "C05A00"
    white  = "FFFFFF"
    light  = "FFF8F3"

    header_font  = Font(name='Arial', bold=True, color=white, size=11)
    header_fill  = PatternFill('solid', start_color=navy)
    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Side(style='thin', color='D0D4E0')
    border = Border(bottom=thin)

    # Title row
    ws.merge_cells('A1:N1')
    title = ws['A1']
    title.value = f"Payroll Report  |  {period_start}  to  {period_end}"
    title.font  = Font(name='Arial', bold=True, color=white, size=13)
    title.fill  = PatternFill('solid', start_color=orange)
    title.alignment = center
    ws.row_dimensions[1].height = 32

    headers = [
        'Employee ID', 'Full Name', 'Position', 'Department',
        'Payment Type', 'Rate', 'Total Hours',
        'Gross Pay', 'Bonus', 'Overtime', 'Deductions',
        'Amount Due', 'Net Paid', 'Status',
    ]
    col_widths = [14, 24, 20, 18, 14, 16, 14, 14, 12, 12, 14, 14, 14, 10]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = PatternFill('solid', start_color=navy)
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[2].height = 26
    ws.freeze_panes = 'A3'

    PESO_COLS = {8, 9, 10, 11, 12, 13}  # Gross, Bonus, OT, Deductions, Due, Net

    for row_idx, r in enumerate(rows, 3):
        fill = PatternFill('solid', start_color='F7F8FB' if row_idx % 2 == 0 else 'FFFFFF')
        status_color = "1a9e7a" if r['status'] == 'Paid' else "c0365a"

        values = [
            r['employee_id'], r['full_name'], r['position'], r['department'],
            r['payment_type'], r['rate_label'], r['total_hours'],
            r['gross_pay'], r['bonus'], r['overtime'], r['deductions'],
            r['amount_due'], r['net_paid'], r['status'],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill   = fill
            cell.border = border
            cell.font   = Font(name='Arial', size=10,
                               color=status_color if col == 14 else '000000',
                               bold=(col == 14))
            if col in PESO_COLS:
                cell.number_format = '₱#,##0.00'
                cell.alignment = right
            elif col == 7:
                cell.number_format = '0.00'
                cell.alignment = right
            else:
                cell.alignment = left

    # Totals row
    if rows:
        total_row = len(rows) + 3
        ws.cell(row=total_row, column=2, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
        ws.cell(row=total_row, column=2).fill = PatternFill('solid', start_color='FFF1E6')

        for col in PESO_COLS:
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=total_row, column=col,
                value=f'=SUM({col_letter}3:{col_letter}{total_row - 1})'
            )
            cell.number_format = '₱#,##0.00'
            cell.font      = Font(name='Arial', bold=True, size=10)
            cell.alignment = right
            cell.fill      = PatternFill('solid', start_color='FFF1E6')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"payroll_{period_start}_to_{period_end}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def create_summary_sheet(ws, payroll_runs):
    """Create summary sheet with run totals."""
    headers = [
        "Run ID", "Period Start", "Period End", "Status", 
        "Total Employees", "Total Gross Pay", 
        "Total Additions",  
        "Total Deductions", 
        "Total Net Pay", "Generated By", "Generated At"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_gross = 0
    total_additions = 0
    total_deductions = 0
    total_net = 0
    unique_employees = set()

    for row_idx, run in enumerate(payroll_runs, 2):
        run_employee_ids = set(entry.worker.id for entry in run.entries.all())
        unique_employees.update(run_employee_ids)
        
        # Calculate additions and deductions for this run
        run_additions = 0
        run_deductions = 0
        for entry in run.entries.all():
            for adj in entry.adjustments.all():
                if adj.type == 'addition':
                    run_additions += float(adj.amount)
                elif adj.type == 'deduction':
                    run_deductions += float(adj.amount)
        
        row_data = [
            run.id,
            run.period_start.strftime("%Y-%m-%d"),
            run.period_end.strftime("%Y-%m-%d"),
            run.status,
            len(run_employee_ids),
            float(run.total_gross_pay),
            run_additions,      # ← Additions
            float(run.total_deductions),
            float(run.total_net_pay),
            run.generated_by.username if run.generated_by else "",
            run.generated_at.strftime("%Y-%m-%d %H:%M") if run.generated_at else "",
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
        
        total_gross += float(run.total_gross_pay)
        total_additions += run_additions
        total_deductions += float(run.total_deductions)
        total_net += float(run.total_net_pay)

    # Totals row with additions
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value="TOTALS").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=len(unique_employees)).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_gross).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_additions).font = Font(bold=True)  # ← Additions
    ws.cell(row=total_row, column=8, value=total_deductions).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=total_net).font = Font(bold=True)

    auto_size_columns(ws)

def create_details_sheet(ws, payroll_runs):
    """Create details sheet with all entries."""
    headers = [
        "Run ID", "Period", "Employee ID", "Employee Name", "Position",
        "Total Hours", "Overtime Hours", "Gross Pay", 
        "Additions",      
        "Deductions",     
        "Net Pay"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, run in enumerate(payroll_runs, 2):
        for entry in run.entries.all():
            # Get adjustments
            adjustments = entry.adjustments.all()
            total_additions = sum(a.amount for a in adjustments if a.type == 'addition')
            total_deductions = sum(a.amount for a in adjustments if a.type == 'deduction')
            
            row_data = [
                run.id,
                f"{run.period_start} to {run.period_end}",
                entry.worker.employee_id,
                entry.worker.full_name,
                entry.worker.position or "",
                float(entry.total_working_hours),
                float(entry.total_overtime_hours),
                float(entry.gross_pay),
                float(total_additions),      
                float(total_deductions),     
                float(entry.net_pay),
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)
            row_idx += 1

    auto_size_columns(ws)

def auto_size_columns(ws):
    """Auto-size columns."""
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)