import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.db.models import Sum
from django.db import transaction
from django.utils.dateparse import parse_date

from .models import User, Wallet, Transaction, Attachment, Log, WalletTransfer, Spend
from .serializers import WalletSerializer, TransactionSerializer, AttachmentSerializer, LogSerializer, WalletTransferSerializer, SpendSerializer
from .constants import EXPENSE_CATEGORIES_DATA

from audit.models import AuditLog

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 1000

class CategoryViewSet(viewsets.ViewSet):
    permission_classes = [permissions.AllowAny]

    def list(self, request):
        return Response(EXPENSE_CATEGORIES_DATA)

class WalletViewSet(viewsets.ModelViewSet):
    queryset = Wallet.objects.all()
    serializer_class = WalletSerializer
    permission_classes = [permissions.AllowAny]
    pagination_class = StandardResultsSetPagination

    def perform_create(self, serializer):
        wallet = serializer.save()
        AuditLog.objects.create(
            user=self.request.user if self.request.user.is_authenticated else None,
            action='CREATE',
            model_name='Wallet',
            object_id=str(wallet.id),
            changes={'new_data': serializer.data},
            description="Created a new wallet."
        )

    def perform_update(self, serializer):
        instance = self.get_object()
        old_data = self.get_serializer(instance).data
        wallet = serializer.save()
        AuditLog.objects.create(
            user=self.request.user if self.request.user.is_authenticated else None,
            action='UPDATE',
            model_name='Wallet',
            object_id=str(wallet.id),
            changes={'old_data': old_data, 'new_data': serializer.data},
            description="Updated a wallet."
        )

    def perform_destroy(self, instance):
        old_data = self.get_serializer(instance).data
        wallet_id = instance.id
        instance.delete()
        AuditLog.objects.create(
            user=self.request.user if self.request.user.is_authenticated else None,
            action='DELETE',
            model_name='Wallet',
            object_id=str(wallet_id),
            changes={'old_data': old_data},
            description="Deleted a wallet."
        )

class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer
    permission_classes = [permissions.AllowAny]
    pagination_class = StandardResultsSetPagination

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else User.objects.first()
        transaction = serializer.save(user=user)
        Log.objects.create(
            user=user,
            action="Create Expense",
            new_data=serializer.data
        )
        AuditLog.objects.create(
            user=user,
            action='CREATE',
            model_name='Transaction',
            object_id=str(transaction.id),
            changes={'new_data': serializer.data},
            description=f"Created a new {transaction.transaction_type} transaction."
        )

    def perform_update(self, serializer):
        instance = self.get_object()
        old_data = self.get_serializer(instance).data
        transaction = serializer.save()
        Log.objects.create(
            user=self.request.user if self.request.user.is_authenticated else User.objects.first(),
            action="Update Expense",
            old_data=old_data,
            new_data=serializer.data
        )
        AuditLog.objects.create(
            user=self.request.user if self.request.user.is_authenticated else None,
            action='UPDATE',
            model_name='Transaction',
            object_id=str(transaction.id),
            changes={'old_data': old_data, 'new_data': serializer.data},
            description=f"Updated a {transaction.transaction_type} transaction."
        )

    def perform_destroy(self, instance):
        old_data = self.get_serializer(instance).data
        transaction_id = instance.id
        transaction_type = instance.transaction_type
        Log.objects.create(
            user=self.request.user if self.request.user.is_authenticated else User.objects.first(),
            action="Delete Expense",
            old_data=old_data
        )
        instance.delete()
        AuditLog.objects.create(
            user=self.request.user if self.request.user.is_authenticated else None,
            action='DELETE',
            model_name='Transaction',
            object_id=str(transaction_id),
            changes={'old_data': old_data},
            description=f"Deleted a {transaction_type} transaction."
        )

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        transactions = self.get_queryset()
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if start_date:
            parsed_start = parse_date(start_date)
            if parsed_start:
                transactions = transactions.filter(transaction_date__gte=parsed_start)
        if end_date:
            parsed_end = parse_date(end_date)
            if parsed_end:
                transactions = transactions.filter(transaction_date__lte=parsed_end)
        expenses = transactions.filter(transaction_type='spend funds')
        income = transactions.filter(transaction_type='add funds')
        total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0.0
        total_income = income.aggregate(Sum('amount'))['amount__sum'] or 0.0
        net_balance = total_income - total_expenses
        expenses_by_category = expenses.exclude(category__isnull=True).values('category').annotate(total=Sum('amount')).order_by('-total')
        
        return Response({
            "total_income": total_income,
            "total_expenses": total_expenses,
            "net_balance": net_balance,
            "expenses_by_category": list(expenses_by_category)
        })

class WalletTransferViewSet(viewsets.ModelViewSet):
    queryset = WalletTransfer.objects.all()
    serializer_class = WalletTransferSerializer
    permission_classes = [permissions.AllowAny]

    def perform_create(self, serializer):
        with transaction.atomic():
            wt = serializer.save()
            user = self.request.user if self.request.user.is_authenticated else User.objects.first()
            txn = Transaction.objects.create(
                transaction_type="move funds",
                user=user,
                wallet=wt.from_wallet,
                transaction_date=wt.created_at.date(),
                amount=wt.amount,
                counterparty=wt.to_wallet.name,
                wallet_transfer=wt,
            )
            wt.transaction = txn
            wt.save(update_fields=["transaction"])

class SpendViewSet(viewsets.ModelViewSet):
    queryset = Spend.objects.all()
    serializer_class = SpendSerializer
    permission_classes = [permissions.AllowAny]

    def perform_create(self, serializer):
        from datetime import datetime as dt
        with transaction.atomic():
            sp = serializer.save()
            user = self.request.user if self.request.user.is_authenticated else User.objects.first()
            tx_date_str = self.request.data.get('transaction_date', '')
            try:
                tx_date = dt.strptime(tx_date_str, '%Y-%m-%d').date() if tx_date_str else sp.created_at.date()
            except (ValueError, TypeError):
                tx_date = sp.created_at.date()
            txn = Transaction.objects.create(
                transaction_type="spend funds",
                user=user,
                wallet=sp.wallet,
                category=sp.category,
                transaction_date=tx_date,
                amount=sp.amount,
                note=self.request.data.get('note', ''),
                counterparty=self.request.data.get('counterparty', ''),
                spend=sp,
            )
            sp.transaction = txn
            sp.save(update_fields=["transaction"])

class AttachmentViewSet(viewsets.ModelViewSet):
    queryset = Attachment.objects.all()
    serializer_class = AttachmentSerializer
    permission_classes = [permissions.AllowAny]

class LogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Log.objects.all()
    serializer_class = LogSerializer
    permission_classes = [permissions.AllowAny]
    pagination_class = StandardResultsSetPagination

class ExportExpensesViewSet(viewsets.ViewSet):
    """
    ViewSet for exporting expenses to XLSX
    """
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        """
        GET /api/expenses/export/xlsx/
        Export expenses for a date range to XLSX.
        Query params: start_date, end_date, category, wallet
        """
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        category = request.query_params.get('category')
        wallet_id = request.query_params.get('wallet_id')

        # Filter transactions (spend funds only)
        transactions = Transaction.objects.filter(
            transaction_type='spend funds'
        ).select_related('wallet', 'user')

        if start_date:
            transactions = transactions.filter(transaction_date__gte=start_date)
        if end_date:
            transactions = transactions.filter(transaction_date__lte=end_date)
        if category:
            transactions = transactions.filter(category=category)
        if wallet_id:
            transactions = transactions.filter(wallet_id=wallet_id)

        transactions = transactions.order_by('-transaction_date')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="left", vertical="center")
        money_alignment = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Date", "Description", "Amount (₱)", "Friendly Category", 
            "BIR Category", "Payment Method", "Wallet", "Counterparty", 
            "Note", "Created By", "Created At"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows
        for row_idx, txn in enumerate(transactions, 2):
            # Get category display
            category_display = dict(Transaction.CATEGORY_CHOICES).get(txn.category, txn.category)
            
            row_data = [
                txn.transaction_date.strftime("%Y-%m-%d"),
                txn.note[:100] if txn.note else "",
                float(txn.amount),
                category_display,
                txn.category.replace('_', ' ').title(),
                txn.transaction_type.replace('_', ' ').title(),
                txn.wallet.name if txn.wallet else "",
                txn.counterparty or "",
                txn.note or "",
                txn.user.username if txn.user else "",
                txn.created_at.strftime("%Y-%m-%d %H:%M") if txn.created_at else "",
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                
                if isinstance(value, (int, float)):
                    cell.alignment = money_alignment
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = cell_alignment

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_length = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_length

        # Add summary row
        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        total_cell = ws.cell(
            row=summary_row, 
            column=3, 
            value=sum(float(t.amount) for t in transactions)
        )
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00'

        # Create response
        filename = f"expenses_{start_date or 'all'}_{end_date or 'all'}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response